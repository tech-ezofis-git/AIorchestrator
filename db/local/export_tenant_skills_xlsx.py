#!/usr/bin/env python3
"""Export tenant SQLite tables to a readable Excel workbook (3 data sheets + README)."""
from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_REPO = Path(__file__).resolve().parents[2]
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))

from app.tenant_skills.store import default_sqlite_path  # noqa: E402

DEFAULT_OUT = Path(__file__).resolve().parent / "tenant_skills_export.xlsx"


def _short(text: object, n: int = 100) -> str:
    if text is None:
        return ""
    cleaned = str(text).replace("\r\n", " ").replace("\n", " ").strip()
    if len(cleaned) <= n:
        return cleaned
    return cleaned[: n - 1] + "…"


def _autosize(ws, max_width: int = 36) -> None:
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        letter = get_column_letter(col[0].column)
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def _style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="E8E4F3")
    header_font = Font(bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def export(db_path: Path, out_path: Path) -> None:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("README", 0)
    for row in (
        ("Tenant skills DB export",),
        (),
        ("Sheet", "What it is"),
        ("Skills", "tenant_skills — custom .md skill uploads (prepended to skill body)."),
        ("Rules", "tenant_rules — custom .mdc rule uploads (appended as rules)."),
        ("Audit Log", "tenant_skill_rule_logs — change history (short summary only, not full bodies)."),
        (),
        ("Note", "Platform defaults (SKILL.md, *.mdc on disk) are read-only and not stored in this DB."),
        ("Source DB", str(db_path)),
    ):
        readme.append(list(row))
    readme["A1"].font = Font(bold=True, size=14)
    readme.column_dimensions["A"].width = 14
    readme.column_dimensions["B"].width = 78

    ws_sk = wb.create_sheet("Skills")
    skill_headers = (
        "id",
        "tenant_id",
        "agent",
        "slug",
        "source_file",
        "is_custom",
        "is_active",
        "body (short)",
    )
    ws_sk.append(skill_headers)
    skill_rows = conn.execute("SELECT * FROM tenant_skills ORDER BY id").fetchall()
    if skill_rows:
        for row in skill_rows:
            ws_sk.append(
                [
                    row["id"],
                    row["tenant_id"],
                    row["agent"],
                    row["slug"],
                    row["source_file"] or "",
                    row["is_custom"],
                    row["is_active"],
                    _short(row["body"]),
                ]
            )
    else:
        ws_sk.append(
            [
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "(no rows — upload a .md file to add custom skills)",
            ]
        )
    _style_header(ws_sk)
    _autosize(ws_sk)

    ws_ru = wb.create_sheet("Rules")
    rule_headers = (
        "id",
        "tenant_id",
        "agent",
        "slug",
        "source_file",
        "is_custom",
        "is_active",
        "always_apply",
        "body (short)",
        "updated_by",
        "updated_at",
    )
    ws_ru.append(rule_headers)
    rule_rows = conn.execute("SELECT * FROM tenant_rules ORDER BY id").fetchall()
    for row in rule_rows:
        ws_ru.append(
            [
                row["id"],
                row["tenant_id"],
                row["agent"],
                row["slug"],
                row["source_file"] or "",
                row["is_custom"],
                row["is_active"],
                row["always_apply"],
                _short(row["body"]),
                row["updated_by"] or "",
                row["updated_at"],
            ]
        )
    if not rule_rows:
        ws_ru.append(["—", "—", "—", "—", "—", "—", "—", "—", "(no custom rules yet)", "—", "—"])
    _style_header(ws_ru)
    _autosize(ws_ru, max_width=42)

    ws_log = wb.create_sheet("Audit Log")
    log_headers = (
        "id",
        "tenant_id",
        "agent",
        "item_type",
        "rule_id",
        "action",
        "summary",
        "changed_by",
        "changed_at",
    )
    ws_log.append(log_headers)
    log_rows = conn.execute("SELECT * FROM tenant_skill_rule_logs ORDER BY id").fetchall()
    for row in log_rows:
        action = row["action"]
        if action in ("ENABLE", "DISABLE"):
            summary = f"Active flag set to {'on' if action == 'ENABLE' else 'off'}"
        elif action == "CREATE":
            summary = "Created: " + _short(row["new_value"], 80)
        else:
            summary = "Updated: " + _short(row["new_value"], 80)
        ws_log.append(
            [
                row["id"],
                row["tenant_id"],
                row["agent"],
                row["item_type"],
                row["item_id"],
                action,
                summary,
                row["changed_by"] or "",
                row["changed_at"],
            ]
        )
    if not log_rows:
        ws_log.append(["—", "—", "—", "—", "—", "—", "(no log entries yet)", "—", "—"])
    _style_header(ws_log)
    _autosize(ws_log, max_width=50)

    conn.close()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved {out_path}")
    print("Sheets:", wb.sheetnames)


def main() -> None:
    export(default_sqlite_path(), DEFAULT_OUT)


if __name__ == "__main__":
    main()
